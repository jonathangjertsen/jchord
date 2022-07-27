"""
Tools for working with chord progressions.
"""
from collections import namedtuple
from typing import Hashable, List, Set, Union

from jchord.knowledge import REPETITION_SYMBOL
from jchord.core import CompositeObject
from jchord.chords import Chord
from jchord.midi import read_midi_file, notes_to_messages, MidiNote
from jchord.group_notes_to_chords import group_notes_to_chords


class InvalidProgression(Exception):
    """Raised when encountering what seems like an invalid chord progression."""


def _string_to_progression(string: str) -> List[Chord]:
    string = string.strip()

    if string == "":
        return []

    progression = []

    for name in string.split():
        name = name.strip()

        if name == REPETITION_SYMBOL:
            if not progression:
                raise InvalidProgression(
                    "Can't repeat before at least one chord has been added"
                )

            progression.append(progression[-1])
        else:
            progression.append(Chord.from_name(name))
    return progression


class MidiConversionSettings(object):
    def __init__(
        self,
        filename: str,
        instrument: int = 1,
        tempo: int = 120,
        beats_per_chord: Union[int, list] = 2,
        velocity: int = 100,
        repeat: str = "replay",
        effect=None,
    ):
        self.filename = filename
        self.instrument = int(instrument)
        self.tempo = int(tempo)
        if isinstance(beats_per_chord, str):
            beats_per_chord = float(beats_per_chord)
        self.beats_per_chord = beats_per_chord
        self.velocity = int(velocity)
        self.repeat = repeat
        self.effect = effect
        self.progression = None

    def set(self, **kwargs):
        for key, value in kwargs.items():
            setattr(self, key, value)


class ChordProgression(CompositeObject):
    """
    Represents a chord progression.

    There are many ways to create a ``ChordProgression`` object.

    **From a string**

    Use the ``from_string`` method to generate a chord progression from a string.

    >>> ChordProgression.from_string("Dm7 -- Gm7 Am7")
    ChordProgression([Chord(name='Dm7', root=Note('D', 4), intervals=Intervals(name='m7', semitones=[0, 3, 7, 10])), Chord(name='Dm7', root=Note('D', 4), intervals=Intervals(name='m7', semitones=[0, 3, 7, 10])), Chord(name='Gm7', root=Note('G', 4), intervals=Intervals(name='m7', semitones=[0, 3, 7, 10])), Chord(name='Am7', root=Note('A', 4), intervals=Intervals(name='m7', semitones=[0, 3, 7, 10]))])

    **From a text file**

    Use the ``from_txt`` method to generate a chord progression from a text file.
    If example.txt contains the text "Am7 D7", then ``ChordProgression.from_txt("example.txt")``
    will produce the same result as ``ChordProgression.from_string("Am7 D7")``.

    **From an Excel file**

    Use the ``from_xlsx`` method to generate a chord progression from an Excel spreadsheet.
    If example.xlsx contains the following cells:

        +-----+-----+
        |  C  |   D |
        +-----+-----+
        | Em7 |  G7 |
        +-----+-----+

    Then the result is equivalent to calling ``ChordProgression.from_string("C D Em7 G7")``.

    .. note::
        This feature requires ``openpyxl``, which you can get with ``pip install openpyxl``.

    **From a MIDI file**

    Use the ``from_midi`` method to generate a chord progression from a MIDI file.
    If ``example.mid`` contains some chords that you would like to convert to a ``ChordProgression``,
    use ``ChordProgression.from_midi("example.mid")``.
    For best results, the MIDI file should contain a single instrument with chords played as straight as possible.

    .. note::
        This feature requires ``mido``, which you can get with ``pip install mido``.
    """

    class _DummyChord(object):
        """Mocks a ChordWithProgression object"""

        def midi(self):
            return []

    DUMMY_CHORD = _DummyChord()

    def __init__(self, progression: List[Chord]):
        self.progression = progression

    def __len__(self):
        return len(self.progression)

    def _keys(self) -> Hashable:
        return (self.progression,)

    @classmethod
    def from_string(cls, string: str) -> "ChordProgression":
        return cls(_string_to_progression(string))

    @classmethod
    def from_txt(cls, filename: str) -> "ChordProgression":
        with open(filename) as file:
            return cls(_string_to_progression(file.read()))

    @classmethod
    def from_xlsx(cls, filename: str) -> "ChordProgression":
        from openpyxl import load_workbook

        workbook = load_workbook(filename)
        sheet = workbook.active
        names = []
        for row in sheet.iter_rows():
            for cell in row:
                name = cell.value
                if not name:
                    name = REPETITION_SYMBOL
                names.append(name)
        return cls.from_string(" ".join(names))

    @classmethod
    def from_midi_file(cls, filename: str) -> "ChordProgression":
        notes = read_midi_file(filename)
        progression = []
        for chord in group_notes_to_chords(notes):
            progression.append(Chord.from_midi([note.note for note in chord]))
        return cls(progression)

    from_midi = from_midi_file

    def chords(self) -> Set[Chord]:
        """
        Returns the set of chords in the progression.

        >>> ChordProgression.from_string("Am7 D7").chords() # doctest: +SKIP
        {Chord(name='D7', root=Note('D', 4), intervals=Intervals(name='7', semitones=[0, 4, 7, 10])), Chord(name='Am7', root=Note('A', 4), intervals=Intervals(name='m7', semitones=[0, 3, 7, 10]))}
        """
        return set(self.progression)

    def midi(self) -> List[List[int]]:
        """
        Returns the MIDI values for each chord in the progression.

        >>> ChordProgression.from_string("Am7 D7").midi()
        [[69, 72, 76, 79], [62, 66, 69, 72]]
        """
        return [chord.midi() for chord in self.progression]

    def transpose(self, shift: int):
        """
        Transposes all chords in the progression by the given shift.

        >>> ChordProgression.from_string("Am7 D7").transpose(2).to_string().strip()
        'Bm7  E7'
        """
        return ChordProgression([chord.transpose(shift) for chord in self.progression])

    def to_string(
        self, chords_per_row: int = 4, column_spacing: int = 2, newline: str = "\n"
    ) -> str:
        """
        Returns the string representation of the chord progression.
        """
        max_len = max(len(chord.name) for chord in self.progression)
        column_width = max_len + column_spacing

        column = 0
        output = []
        prev_chord = None
        for chord in self.progression:
            if prev_chord == chord:
                chord_name = REPETITION_SYMBOL
            else:
                chord_name = chord.name

            output.append(chord_name)
            output.append(" " * (column_width - len(chord_name)))

            column += 1
            if column % chords_per_row == 0:
                column = 0
                output.append(newline)

            prev_chord = chord

        return "".join(output) + newline

    def to_txt(
        self,
        filename: str,
        chords_per_row: int = 4,
        column_spacing: int = 2,
        newline: str = "\n",
    ):
        """
        Saves the string representation of the chord progression to a text file.
        """
        output_str = self.to_string(
            chords_per_row=chords_per_row,
            column_spacing=column_spacing,
            newline=newline,
        )
        with open(filename, "w") as file:
            file.write(output_str)

    def to_xlsx(self, filename: str, chords_per_row: int = 4):
        """
        Saves the chord progression to an Excel file.

        .. note::
            This feature requires ``openpyxl``, which you can get with ``pip install openpyxl``.
        """
        from openpyxl import Workbook

        workbook = Workbook()
        worksheet = workbook.active

        row = 1
        column = 1
        prev_chord = None
        for chord in self.progression:
            if prev_chord == chord:
                chord_name = REPETITION_SYMBOL
            else:
                chord_name = chord.name

            worksheet.cell(row=row, column=column).value = chord_name

            column += 1
            if (column - 1) % chords_per_row == 0:
                column = 1
                row += 1

            prev_chord = chord

        workbook.save(filename)

    def to_pdf(self, filename, **kwargs):
        """
        Creates a PDF.

        .. note::
            This feature requires ``reportlab``, which you can get with ``pip install reportlab``.
        """
        song = Song([SongSection(filename, self)])
        return song.to_pdf(filename, **kwargs)

    def to_midi(self, settings: MidiConversionSettings, **kwargs):
        """
        Saves the chord progression to a MIDI file.

        .. note::
            This feature requires ``mido``, which you can get with ``pip install mido``.
        """
        if not isinstance(settings, MidiConversionSettings) or kwargs:
            raise ValueError(
                "to_midi now takes a MidiConversionSettings object, not individual arguments; see README.md"
            )

        repeat_options = {"replay", "hold"}
        assert (
            settings.repeat in repeat_options
        ), f"repeat argument must be one of: {repeat_options}"

        import mido

        mid = mido.MidiFile()
        track = mido.MidiTrack()
        mid.tracks.append(track)

        # Ensure beats_per_chord is a list
        if isinstance(settings.beats_per_chord, (int, float)):
            settings.beats_per_chord = [
                settings.beats_per_chord for _ in range(len(self.progression))
            ]
        assert len(settings.beats_per_chord) == len(
            self.progression
        ), "len(settings.beats_per_chord) is {}, which is not equal to the number of chords in the progression ({})".format(
            len(settings.beats_per_chord), len(self.progression)
        )

        seconds_per_chord = [
            (60 / settings.tempo) * bpc for bpc in settings.beats_per_chord
        ]
        ticks_per_chord = [
            int(
                mido.second2tick(
                    spc, mid.ticks_per_beat, mido.bpm2tempo(settings.tempo)
                )
            )
            for spc in seconds_per_chord
        ]
        track.append(
            mido.MetaMessage("set_tempo", tempo=mido.bpm2tempo(settings.tempo))
        )
        track.append(mido.Message("program_change", program=settings.instrument))

        played_chords = []
        prev_chord = None
        time = 0

        for chord, tpc in zip(self.midi(), ticks_per_chord):
            if chord == prev_chord and settings.repeat == "hold":
                played_chords[-1] = [
                    pnote._replace(duration=pnote.duration + tpc)
                    for pnote in played_chords[-1]
                ]
            else:
                played_chords.append(
                    [
                        MidiNote(
                            note=note,
                            velocity=settings.velocity,
                            time=time,
                            duration=tpc,
                        )
                        for note in chord
                    ]
                )
            prev_chord = chord
            time += tpc

        settings.set(progression=self)
        settings.set(played_chords=played_chords)
        settings.set(midi_track=track)

        if settings.effect:
            settings.effect.set_settings(settings)
            played_chords = [settings.effect.apply(chord) for chord in played_chords]

        played_notes = [note for chord in played_chords for note in chord]
        for message in notes_to_messages(played_notes, velocity=settings.velocity):
            track.append(message)
        mid.save(settings.filename)


SongSection = namedtuple("SongSection", "name, progression")
SongSection.__doc__ = """Represents a section in a Song."""


class Song(CompositeObject):
    """Represents a song (a series of sections)."""

    def __init__(self, sections: List[SongSection]):
        self.sections = sections

    def _keys(self):
        return (self.sections,)

    def to_string(
        self, chords_per_row: int = 4, column_spacing: int = 2, newline: str = "\n"
    ):
        """Returns the string representation of the song."""
        out = []
        multiplier = 1
        for i, section in enumerate(self.sections):
            if multiplier > 1:
                multiplier -= 1
                continue

            for j in range(i + 1, len(self.sections)):
                if self.sections[j] is self.sections[i]:
                    multiplier += 1
                else:
                    break

            if multiplier > 1:
                section_name = f"{section.name} (x{multiplier})"
            else:
                section_name = section.name

            out.append(f"{section_name}{newline}{'=' * len(section_name)}{newline}")
            out.append(
                section.progression.to_string(
                    chords_per_row=chords_per_row,
                    column_spacing=column_spacing,
                    newline=newline,
                )
            )
            out.append(newline)
        combined = "".join(out)
        combined = combined.replace(3 * newline, 2 * newline)
        combined = combined.strip() + newline
        combined = newline.join(line.strip() for line in combined.split(newline))
        return combined

    def to_pdf(
        self,
        filename,
        *,
        font="Helvetica",
        fontsize=16,
        chords_per_row: int = 4,
        margin_factor_w=1.0,
        margin_factor_h=1.0,
        spacing_factor_w=1.0,
        spacing_factor_h=1.0,
        spacing_min_h=75,
    ):
        """
        Creates a PDF.

        .. note::
            This feature requires ``reportlab``, which you can get with ``pip install reportlab``.
        """
        from reportlab.pdfgen.canvas import Canvas

        fontsize = float(fontsize)
        chords_per_row = int(chords_per_row)
        margin_factor_w = float(margin_factor_w)
        margin_factor_h = float(margin_factor_h)
        spacing_factor_w = float(spacing_factor_w)
        spacing_factor_h = float(spacing_factor_h)
        spacing_min_h = float(spacing_min_h)

        canvas = Canvas(filename)
        canvas.setFont(font, fontsize)
        WIDTH, HEIGHT = canvas._pagesize
        n_rows = (
            sum(len(section.progression) for section in self.sections) / chords_per_row
        )

        SPACING_BASE_W = 600
        SPACING_BASE_H = 750

        spacing_w = spacing_factor_w * (
            SPACING_BASE_W / (chords_per_row + 2 * margin_factor_w)
        )
        spacing_h = min(spacing_factor_h * SPACING_BASE_H / n_rows, spacing_min_h)

        column = 0
        row = 0
        prev_chord = None
        for name, progression in self.sections:
            column += 1
            canvas.drawString(
                spacing_w * (row + margin_factor_w),
                HEIGHT - spacing_h * (column + margin_factor_h),
                name,
            )
            column += 1

            for chord in progression.progression:
                if prev_chord == chord:
                    chord_name = REPETITION_SYMBOL
                else:
                    chord_name = chord.name
                canvas.drawString(
                    spacing_w * (row + margin_factor_w),
                    HEIGHT - spacing_h * (column + margin_factor_h),
                    chord_name,
                )
                row += 1
                if row % chords_per_row == 0:
                    row = 0
                    column += 1
                prev_chord = chord

        canvas.save()
